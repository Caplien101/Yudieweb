"""
pdf_to_xlsx.py
==============
Converter PDF -> XLSX yang "powerful" tapi tetap simpel untuk di-pakai
sebagai modul (bisa dipanggil dari FastAPI/Flask) maupun dari CLI.

Strategi (pipeline):
1. Buka PDF, iterasi tiap halaman.
2. Cek apakah halaman punya layer teks asli:
   - Ada teks       -> coba ekstrak tabel dengan pdfplumber (mode "lines"
                        lalu fallback "text"), lalu fallback ke camelot
                        (lattice -> stream) kalau pdfplumber gagal/kosong.
   - Tidak ada teks  -> anggap halaman hasil scan -> render ke gambar
                        lalu OCR pakai pytesseract, hasilnya di-parse
                        jadi baris/kolom sederhana.
3. Setiap tabel yang berhasil dideteksi ditulis ke sheet Excel terpisah
   (Sheet1_p1_t1, dst), dengan header di-bold dan kolom di-auto-fit.
4. Semua proses dibungkus try/except per halaman supaya 1 halaman error
   tidak menggagalkan seluruh dokumen -> hasil tetap "graceful".

Dependency yang dipakai (semua sudah lazim & ringan):
    pdfplumber, camelot-py, openpyxl, pytesseract (+ binary tesseract)

Instalasi (kalau belum ada):
    pip install pdfplumber camelot-py[cv] openpyxl pytesseract pillow --break-system-packages
    apt-get install -y tesseract-ocr        # untuk binary OCR-nya
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("pdf_to_xlsx")


# --------------------------------------------------------------------------- #
# Struktur data
# --------------------------------------------------------------------------- #

@dataclass
class ExtractedTable:
    page: int
    index: int
    rows: list[list[str]]
    source: str  # "pdfplumber" | "camelot" | "ocr"


@dataclass
class ConversionReport:
    total_pages: int = 0
    tables_found: int = 0
    ocr_pages: int = 0
    failed_pages: list[int] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Ekstraksi halaman bertext (native)
# --------------------------------------------------------------------------- #

def _clean_cell(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _extract_with_pdfplumber(page: "pdfplumber.page.Page") -> list[list[list[str]]]:
    """Coba dua strategi deteksi tabel: garis eksplisit, lalu berbasis teks."""
    strategies = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {"vertical_strategy": "text", "horizontal_strategy": "text"},
    ]
    for strat in strategies:
        try:
            tables = page.extract_tables(table_settings=strat)
        except Exception as exc:  # pragma: no cover
            logger.debug("pdfplumber strategy %s gagal: %s", strat, exc)
            continue
        if tables:
            cleaned = [
                [[_clean_cell(c) for c in row] for row in table]
                for table in tables
                if table and any(any(_clean_cell(c) for c in row) for row in table)
            ]
            if cleaned:
                return cleaned
    return []


def _extract_with_camelot(pdf_path: str, page_number: int) -> list[list[list[str]]]:
    """Fallback ke camelot untuk tabel yang lebih kompleks/tanpa garis rapi."""
    try:
        import camelot
    except ImportError:
        logger.warning("camelot tidak terpasang, lewati fallback ini.")
        return []

    for flavor in ("lattice", "stream"):
        try:
            tables = camelot.read_pdf(
                pdf_path, pages=str(page_number), flavor=flavor
            )
        except Exception as exc:  # pragma: no cover
            logger.debug("camelot flavor=%s gagal di hal %s: %s", flavor, page_number, exc)
            continue
        if tables.n > 0:
            result = []
            for t in tables:
                df = t.df
                rows = [[_clean_cell(v) for v in row] for row in df.values.tolist()]
                if any(any(r) for r in rows):
                    result.append(rows)
            if result:
                return result
    return []


# --------------------------------------------------------------------------- #
# OCR untuk halaman hasil scan
# --------------------------------------------------------------------------- #

def _extract_with_ocr(page: "pdfplumber.page.Page") -> list[list[str]]:
    """
    OCR sederhana: render halaman jadi gambar lalu baca teks per baris.
    Kolom dipisah berdasar >=2 spasi berurutan (heuristik, cukup untuk
    tabel yang cukup rapi). Untuk scan kompleks, pertimbangkan layanan
    khusus table-recognition (mis. PaddleOCR table module / Document AI).
    """
    try:
        import pytesseract
    except ImportError:
        logger.warning("pytesseract tidak terpasang, halaman scan dilewati.")
        return []

    try:
        image = page.to_image(resolution=300).original
        text = pytesseract.image_to_string(image)
    except Exception as exc:
        logger.warning("OCR gagal di halaman: %s", exc)
        return []

    rows = []
    for line in text.splitlines():
        line = line.rstrip()
        if not line.strip():
            continue
        cols = re.split(r"\s{2,}", line.strip())
        rows.append([_clean_cell(c) for c in cols])
    return rows


# --------------------------------------------------------------------------- #
# Penulisan ke XLSX
# --------------------------------------------------------------------------- #

def _write_sheet(wb: Workbook, name: str, rows: list[list[str]]) -> None:
    # Nama sheet Excel max 31 karakter & tak boleh karakter tertentu
    safe_name = re.sub(r"[\\/*?:\[\]]", "_", name)[:31]
    ws: Worksheet = wb.create_sheet(title=safe_name)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    # Auto-fit lebar kolom (perkiraan berbasis panjang teks terpanjang)
    max_cols = max((len(r) for r in rows), default=0)
    for c_idx in range(1, max_cols + 1):
        longest = max(
            (len(str(row[c_idx - 1])) for row in rows if c_idx - 1 < len(row)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(longest + 2, 10), 60)


# --------------------------------------------------------------------------- #
# Fungsi utama (bisa dipanggil dari kode lain / FastAPI endpoint)
# --------------------------------------------------------------------------- #

def convert_pdf_to_xlsx(
    pdf_path: str,
    xlsx_path: str,
    ocr_fallback: bool = True,
) -> ConversionReport:
    """
    Konversi satu file PDF menjadi satu file XLSX (multi-sheet, satu
    sheet per tabel yang terdeteksi). Mengembalikan ConversionReport
    berisi ringkasan proses (dipakai untuk logging / respons API).
    """
    report = ConversionReport()
    wb = Workbook()
    wb.remove(wb.active)  # hapus sheet kosong default

    with pdfplumber.open(pdf_path) as pdf:
        report.total_pages = len(pdf.pages)

        for page_number, page in enumerate(pdf.pages, start=1):
            try:
                has_text = bool((page.extract_text() or "").strip())
                tables: list[list[list[str]]] = []

                if has_text:
                    tables = _extract_with_pdfplumber(page)
                    if not tables:
                        tables = _extract_with_camelot(pdf_path, page_number)

                if not tables and ocr_fallback:
                    ocr_rows = _extract_with_ocr(page)
                    if ocr_rows:
                        tables = [ocr_rows]
                        report.ocr_pages += 1

                if not tables:
                    logger.info("Halaman %s: tidak ada tabel terdeteksi.", page_number)
                    continue

                for t_idx, rows in enumerate(tables, start=1):
                    sheet_name = f"p{page_number}_t{t_idx}"
                    _write_sheet(wb, sheet_name, rows)
                    report.tables_found += 1

                logger.info("Halaman %s: %s tabel ditulis.", page_number, len(tables))

            except Exception as exc:
                logger.error("Halaman %s gagal diproses: %s", page_number, exc)
                report.failed_pages.append(page_number)

    if not wb.sheetnames:
        # Tidak ada tabel sama sekali -> tetap hasilkan file dengan pesan
        ws = wb.create_sheet("Info")
        ws["A1"] = "Tidak ada tabel yang berhasil terdeteksi di PDF ini."

    wb.save(xlsx_path)
    logger.info(
        "Selesai. %s tabel, %s halaman OCR, %s halaman gagal -> %s",
        report.tables_found, report.ocr_pages, len(report.failed_pages), xlsx_path,
    )
    return report


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main() -> None:
    parser = argparse.ArgumentParser(description="Konversi PDF ke XLSX (auto tabel + OCR fallback).")
    parser.add_argument("input", help="Path file PDF sumber")
    parser.add_argument("output", nargs="?", help="Path file XLSX tujuan (default: sama nama, .xlsx)")
    parser.add_argument("--no-ocr", action="store_true", help="Matikan fallback OCR untuk halaman scan")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        logger.error("File tidak ditemukan: %s", input_path)
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path.with_suffix(".xlsx")

    report = convert_pdf_to_xlsx(
        str(input_path), str(output_path), ocr_fallback=not args.no_ocr
    )

    print(f"\nSelesai -> {output_path}")
    print(f"Total halaman : {report.total_pages}")
    print(f"Tabel ditemukan: {report.tables_found}")
    print(f"Halaman via OCR: {report.ocr_pages}")
    if report.failed_pages:
        print(f"Halaman gagal  : {report.failed_pages}")


if __name__ == "__main__":
    main()
